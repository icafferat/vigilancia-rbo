from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
from fastapi import FastAPI, Depends, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware
import pandas as pd
import io
from datetime import datetime

# Importes de tus archivos locales
import database
import models
from database import engine, get_db
USER_ADMIN = "admin"
PASS_ADMIN = "aeronautica2025"

# Crear tablas si no existen
models.Base.metadata.create_all(bind=engine)

app = FastAPI()
# Necesario para el Login
app.add_middleware(SessionMiddleware, secret_key="rbo_2026_secret_key")

# --- LÓGICA DE CÁLCULO RBO Y CALENDARIO ---
# Definirla aquí arriba evita errores de "function not defined"
def calcular_perfil_rbo(prob, sev, acft, vuelos, estaciones):
    riesgo_sms = (prob or 0) * (sev or 0)
    # Factor de exposición: basado en tamaño de flota y operaciones
    exp = ((acft or 0) / 5) + ((vuelos or 0) / 200) + ((estaciones or 0) / 2)
    total = riesgo_sms + exp

    if total > 35:
        prioridad, color = "Muy Alta", "#e74c3c"
        cronograma = "Ene: BASE PRINCIPAL | Mar: PLATAFORMA | May: RUTA | Jul: ESTACIÓN | Sep: REGISTROS | Nov: CABINA"
    elif total > 18:
        prioridad, color = "Media", "#f39c12"
        cronograma = "Feb: BASE PRINCIPAL | Jun: PLATAFORMA | Oct: IDE"
    else:
        prioridad, color = "Baja", "#27ae60"
        cronograma = "Jun: PLATAFORMA o RUTA"
        
    return prioridad, color, cronograma

# --- RUTAS DE ACCESO ---
@app.get("/login", response_class=HTMLResponse)
async def login_page():
    return """
    <html><head><title>RBO Login</title><style>body{font-family:sans-serif;background:#1e3a5f;display:flex;justify-content:center;align-items:center;height:100vh;margin:0;}.card{background:white;padding:30px;border-radius:10px;width:320px;text-align:center;}input{width:100%;margin:10px 0;padding:12px;border:1px solid #ddd;border-radius:5px;}button{width:100%;padding:12px;background:#3498db;color:white;border:none;border-radius:5px;cursor:pointer;}</style></head>
    <body><div class="card"><h2>SISTEMA RBO 2026</h2><form action="/login" method="post"><input type="text" name="username" placeholder="Usuario" required><input type="password" name="password" placeholder="Contraseña" required><button type="submit">INGRESAR</button></form></div></body></html>
    """

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    # Agregamos comillas y los dos puntos finales
    if username == USER_ADMIN and password == PASS_ADMIN:
        request.session["user"] = username
        return RedirectResponse(url="/", status_code=303)
    
    # Respuesta en caso de error
    return HTMLResponse("<script>alert('Usuario o contraseña incorrectos'); window.location='/login';</script>")
@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login")

# --- DASHBOARD PRINCIPAL ---
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"): 
        return RedirectResponse(url="/login")

    ops = db.query(models.Operador).all()
    
    conteos = {"Muy Alta": 0, "Media": 0, "Baja": 0}
    nombres_labels = []
    puntajes_sms = []
    
    filas = ""
    for o in ops:
        # 1. Ejecutar lógica RBO
        prioridad, color, cronograma = calcular_perfil_rbo(
            o.probabilidad, o.severidad, o.aeronaves or 0, o.vuelos_mes or 0, o.estaciones or 0
        )
        
        # 2. Llenar datos para gráficas
        conteos[prioridad] += 1
        nombres_labels.append(o.nombre)
        puntajes_sms.append((o.probabilidad or 0) * (o.severidad or 0))
        
        # 3. Construir filas de la tabla
        filas += f"""
        <tr style='border-bottom: 1px solid #eee;'>
            <td style='padding:15px;'><strong>{o.nombre}</strong></td>
            <td style='text-align:center;'>
                <span style='background:{color}; color:white; padding:5px 10px; border-radius:15px; font-size:12px; font-weight:bold;'>
                    {prioridad}
                </span>
            </td>
            <td>
                <div style='background:#f8f9fa; border-left:4px solid {color}; padding:10px; border-radius:4px; font-family:monospace; font-size:13px;'>
                    {cronograma}
                </div>
            </td>
            <td style='text-align:center;'><a href='/eliminar/{o.id}' style='text-decoration:none;'>🗑️</a></td>
        </tr>
        """

    # IMPORTANTE: Aquí debe ir el return f"""
    return f"""
    <html>
    <head>
        <title>RBO 2026</title>
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <style>
            body {{ font-family: sans-serif; background:#f4f7f9; margin:0; }}
            .header {{ background:#1a2a3a; color:white; padding:15px 40px; display:flex; justify-content:space-between; align-items:center; }}
            .container {{ max-width:1100px; margin:20px auto; background:white; padding:25px; border-radius:10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .form-section {{ background:#eef2f7; padding:20px; border-radius:8px; margin-bottom:20px; }}
            .grid-form {{ display:grid; grid-template-columns: repeat(3, 1fr); gap:10px; }}
            input, button {{ padding:10px; border-radius:5px; border:1px solid #ccc; }}
            button {{ background:#1a2a3a; color:white; cursor:pointer; font-weight:bold; }}
            table {{ width:100%; border-collapse:collapse; margin-top:20px; }}
            th {{ background:#f1f4f7; padding:12px; text-align:left; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>VIGILANCIA RBO 2026</h2>
            <div>
                <a href="/exportar" style="color:#2ecc71; text-decoration:none; margin-right:20px; font-weight:bold;">EXCEL 📥</a>
                <a href="/logout" style="color:white; text-decoration:none;">Cerrar</a>
            </div>
        </div>
        <div class="container">
            <details style="background:#fff3cd; border:1px solid #ffeeba; padding:15px; border-radius:8px; margin-bottom:20px; color:#856404;">
                <summary style="cursor:pointer; font-weight:bold;"> Ver Guía de Puntajes (Matriz de Riesgo OACI)</summary>
                <div style="display:grid; grid-template-columns: 1fr 1fr; gap:20px; margin-top:10px; font-size:13px;">
                    <div>
                        <strong>PROBABILIDAD:</strong><br>
                        5: Frecuente (Varias veces al mes)<br>
                        4: Ocasional (Varias veces al año)<br>
                        3: Remota (Poco probable)<br>
                        2: Improbable (No se espera que ocurra)<br>
                        1: Extremadamente improbable
                    </div>
                    <div>
                        <strong>SEVERIDAD:</strong><br>
                        5: Catastrófica (Pérdida de aeronave)<br>
                        4: Peligrosa (Reducción de seguridad)<br>
                        3: Mayor (Incidente serio)<br>
                        2: Menor (Incidente leve)<br>
                        1: Insignificante
                    </div>
                </div>
            </details>

            <div class="form-section">
                <h3>Nuevo Registro de Operador</h3>
                <form action="/registrar" method="post" class="grid-form">
                    <input type="text" name="inspector" placeholder="Inspector Responsable" required style="grid-column: span 1;">
                    <input type="text" name="nombre" placeholder="Nombre Operador (Ej. LATAM, SKY)" required style="grid-column: span 2;">
                    <input type="date" name="fecha" value="{datetime.now().strftime('%Y-%m-%d')}" required>
                    <input type="number" name="probabilidad" placeholder="Prob. (1-5)" min="1" max="5" required title="1: Raro - 5: Frecuente">
                    <input type="number" name="severidad" placeholder="Sev. (1-5)" min="1" max="5" required title="1: Leve - 5: Catastrófico">
                    <input type="number" name="aeronaves" placeholder="N° Aeronaves" required>
                    <input type="number" name="vuelos_mes" placeholder="Vuelos al Mes" required>
                    <input type="number" name="estaciones" placeholder="N° Estaciones" required title="Bases o destinos principales">
                    <button type="submit" style="grid-column: span 3;">GENERAR PLAN DE VIGILANCIA</button>
                </form>
            </div>

            <table>
                <thead><tr><th>Operador</th><th style="text-align:center;">Prioridad</th><th>Cronograma 2026</th><th style="text-align:center;">Acción</th></tr></thead>
                <tbody>{filas}</tbody>
            </table>

            <div style="display:grid; grid-template-columns: 1fr 1fr; gap:20px; margin-top:30px;">
                <div><canvas id="pieChart"></canvas></div>
                <div><canvas id="barChart"></canvas></div>
            </div>
        </div>
        <script>
            new Chart(document.getElementById('pieChart'), {{ 
                type: 'doughnut', 
                data: {{ 
                    labels: ['Muy Alta', 'Media', 'Baja'], 
                    datasets: [{{ 
                        data: [{conteos['Muy Alta']}, {conteos['Media']}, {conteos['Baja']}], 
                        backgroundColor: ['#e74c3c', '#f39c12', '#27ae60'] 
                    }}] 
                }},
                options: {{ plugins: {{ title: {{ display: true, text: 'Distribución de Riesgo' }} }} }}
            }});

            new Chart(document.getElementById('barChart'), {{ 
                type: 'bar', 
                data: {{ 
                    labels: {nombres_labels}, 
                    datasets: [{{ 
                        label: 'Riesgo SMS (P x S)', 
                        data: {puntajes_sms}, 
                        backgroundColor: '#3498db' 
                    }}] 
                }},
                options: {{ plugins: {{ title: {{ display: true, text: 'Puntaje SMS por Operador' }} }} }}
            }});
        </script>
    </body>
    </html>
    """
# --- ACCIONES ---
@app.post("/registrar")
async def registrar(nombre: str = Form(...), probabilidad: int = Form(...), 
                    severidad: int = Form(...), aeronaves: int = Form(...), 
                    vuelos_mes: int = Form(...), estaciones: int = Form(...),
                    fecha: str = Form(...), inspector: str = Form(...), # <-- Nuevo campo
                    db: Session = Depends(get_db)):
    nuevo = models.Operador(
        nombre=nombre, 
        probabilidad=probabilidad,
        severidad=severidad,
        aeronaves=aeronaves,
        vuelos_mes=vuelos_mes,
        estaciones=estaciones,
        inspector=inspector, # <-- Guardar inspector
        fecha=datetime.strptime(fecha, "%Y-%m-%d")
    )
    db.add(nuevo)
    db.commit()
    return RedirectResponse(url="/", status_code=303)

@app.get("/eliminar/{id}")
async def eliminar(id: int, db: Session = Depends(get_db)):
    op = db.query(models.Operador).filter(models.Operador.id == id).first()
    if op: db.delete(op); db.commit()
    return RedirectResponse(url="/", status_code=303)

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image # Aunque no la usemos aquí, es útil tenerla para imágenes
from openpyxl.styles import PatternFill, Font
import pandas as pd
import io
from datetime import datetime
from collections import Counter # Para contar las prioridades

@app.get("/exportar")
async def exportar_excel(db: Session = Depends(get_db)):
    ops = db.query(models.Operador).all()
    
    data_plan = []
    prioridades_list = [] # Para el gráfico
    for o in ops:
        prioridad, color, cronograma = calcular_perfil_rbo(
            o.probabilidad, o.severidad, o.aeronaves or 0, o.vuelos_mes or 0, o.estaciones or 0
        )
        data_plan.append({
            "Operador": o.nombre,
            "Inspector Asignado": getattr(o, 'inspector', 'No asignado'),
            "Fecha Eval": o.fecha.strftime('%d/%m/%Y') if o.fecha else "N/A",
            "Riesgo SMS": (o.probabilidad or 0) * (o.severidad or 0),
            "ACFT": o.aeronaves,
            "Vuelos/Mes": o.vuelos_mes,
            "Prioridad": prioridad,
            "Cronograma 2026": cronograma
        })
        prioridades_list.append(prioridad)

    # Contar las prioridades para el gráfico
    conteo_prioridades = Counter(prioridades_list)
    df_conteo = pd.DataFrame(conteo_prioridades.items(), columns=['Prioridad', 'Cantidad'])
    # Aseguramos que las tres categorías existan para el gráfico
    for p in ["Muy Alta", "Media", "Baja"]:
        if p not in df_conteo['Prioridad'].values:
            df_conteo = pd.concat([df_conteo, pd.DataFrame([{'Prioridad': p, 'Cantidad': 0}])], ignore_index=True)
    df_conteo = df_conteo.sort_values(by='Prioridad', key=lambda x: x.map({"Muy Alta":0, "Media":1, "Baja":2})) # Ordenar

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- Hoja 1: Plan de Vigilancia ---
        df_plan = pd.DataFrame(data_plan)
        df_plan.to_excel(writer, index=False, sheet_name='Plan de Vigilancia')
        
        workbook = writer.book
        worksheet_plan = writer.sheets['Plan de Vigilancia']
        
        fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_naranja = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        font_negrita = Font(bold=True)

        for row in range(2, len(data_plan) + 2):
            celda = worksheet_plan.cell(row=row, column=7) # Columna G es Prioridad
            valor = celda.value
            if valor == "Muy Alta": celda.fill = fill_rojo
            elif valor == "Media": celda.fill = fill_naranja
            elif valor == "Baja": celda.fill = fill_verde
            celda.font = font_negrita

        # --- Hoja 2: Matriz de Referencia ---
        pd.DataFrame([
            {"Nivel": 5, "Probabilidad": "Frecuente", "Severidad": "Catastrófica", "Definición": "Ocurre frecuentemente / Pérdida total"},
            {"Nivel": 4, "Probabilidad": "Ocasional", "Severidad": "Peligrosa", "Definición": "Ocurre a veces / Gran reducción de seguridad"},
            {"Nivel": 3, "Probabilidad": "Remota", "Severidad": "Mayor", "Definición": "Poco probable / Incidente serio"},
            {"Nivel": 2, "Probabilidad": "Improbable", "Severidad": "Menor", "Definición": "Muy poco probable / Incidente leve"},
            {"Nivel": 1, "Probabilidad": "Ext. Improbable", "Severidad": "Insignificante", "Definición": "Casi imposible / Sin consecuencias"},
        ]).to_excel(writer, index=False, sheet_name='Criterios')
        
        # --- Hoja 3: Gráfico de Distribución ---
        worksheet_chart = workbook.create_sheet(title='Distribución RBO')
        df_conteo.to_excel(writer, sheet_name='Distribución RBO', startrow=0, startcol=0, index=False)

        # Crear el gráfico de pastel
        chart = PieChart()
        labels = Reference(worksheet_chart, min_col=1, min_row=2, max_row=len(df_conteo) + 1)
        data = Reference(worksheet_chart, min_col=2, min_row=1, max_row=len(df_conteo) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Distribución de Prioridades de Vigilancia RBO"

        # Colores personalizados para las tajadas (muy similar a tu web)
        for i, (prioridad, cantidad) in enumerate(df_conteo.values):
            if prioridad == "Muy Alta":
                chart.series[0].data_points[i].graphicalProperties.solidFill = "FF0000" # Rojo
            elif prioridad == "Media":
                chart.series[0].data_points[i].graphicalProperties.solidFill = "FFA500" # Naranja
            elif prioridad == "Baja":
                chart.series[0].data_points[i].graphicalProperties.solidFill = "008000" # Verde
        
        chart.width = 15 # cm
        chart.height = 10 # cm
        
        worksheet_chart.add_chart(chart, "D1") # Posicionar el gráfico en la celda D1

    output.seek(0)
    fecha_hoy = datetime.now().strftime("%d_%m_%Y")
    headers = {'Content-Disposition': f'attachment; filename="Plan_RBO_{fecha_hoy}.xlsx"'}
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')